from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, ScatterChart, Reference, Series
import pandas as pd
import os

# Ensure the output directory exists
output_dir = "outputs"
os.makedirs(output_dir, exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = "MMA Stats"

df = pd.read_excel("outputs/processed_stats.xlsx")

# Write data
for row in df[["Name", "Total Attendance", "Sparring Wins", "Sparring Losses", "Efficiency Score"]].itertuples(index=False):
    ws.append(row)

# Add headers manually
for index, column in enumerate(["Name", "Total Attendance", "Sparring Wins", "Sparring Losses", "Efficiency Score"]):
    ws.cell(row=1, column=index + 1, value=column)

# BarChart
bar_chart = BarChart()
bar_chart.title = "Total Attendance by Member"
data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(cats)
ws.add_chart(bar_chart, "G2")

# PieChart
pie_chart = PieChart()
pie_chart.title = "Sparring Wins Distribution"
data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
pie_chart.add_data(data, titles_from_data=True)
pie_chart.set_categories(labels)
ws.add_chart(pie_chart, "G20")

# ScatterChart
scatter_chart = ScatterChart()
scatter_chart.title = "Wins vs. Losses"
xvalues = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)
yvalues = Reference(ws, min_col=4, min_row=2, max_row=ws.max_row)
series = Series(yvalues, xvalues, title="Members")
scatter_chart.series.append(series)
ws.add_chart(scatter_chart, "G38")

wb.save("outputs/MMA_Club_Charts.xlsx")
print("Charts saved successfully in 'outputs/MMA_Club_Charts.xlsx'")